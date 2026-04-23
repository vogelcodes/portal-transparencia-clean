"""Render ARP export bundle to XLSX, CSV, and ODS."""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from odf.opendocument import OpenDocumentSpreadsheet
from odf.style import Style, TextProperties, TableCellProperties
from odf.text import P
from odf.table import Table, TableColumn, TableRow, TableCell


# ── Shared helpers ─────────────────────────────────────────────────────────────

def _v(v):
    return '' if v is None else str(v)


def _brl(v):
    if v is None or v == '':
        return ''
    try:
        return f"R$ {float(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except Exception:
        return _v(v)


def _pct(v):
    return f'{v:.1f}%' if v is not None else ''


# ── XLSX ───────────────────────────────────────────────────────────────────────

def _hdr_font(): return Font(bold=True, color='ffffff', size=11)
def _hdr_fill(): return PatternFill('solid', fgColor='1a3a5c')
def _hdr_align(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def _hdr_border(): return Border(bottom=Side(style='thin', color='ffffff'))


def _write_header(ws, cols):
    for i, (label, *_) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = _hdr_font()
        c.fill = _hdr_fill()
        c.alignment = _hdr_align()
        c.border = _hdr_border()


def _auto_width(ws, min_w=10, max_w=60):
    for col in ws.columns:
        mx = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx + 2, min_w), max_w)


def _sheet_resumo_xlsx(wb, meta):
    ws = wb.create_sheet('Resumo')
    ws['A1'] = 'Relatório ARP - UASG'
    ws['A1'].font = Font(bold=True, size=14)
    rows = [
        ('Código UASG', meta['codigo_uasg']),
        ('Nome', meta['nome_uasg']),
        ('UF / Município', f"{meta['sigla_uf']} — {meta['municipio']}".strip(' —')),
        ('CNPJ', meta['cnpj']),
        ('Gerado em', meta['generated_at']),
        ('Total ARPs', meta['total_arps']),
        ('Total Itens', meta['total_itens']),
        ('Total Saldos', meta['total_saldos']),
    ]
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50


def _sheet_arps_xlsx(wb, arps):
    ws = wb.create_sheet('ARPs')
    cols = [
        ('Nº ATA', 'numero_ata'),
        ('Nº Controle PNCP', 'numero_controle'),
        ('Vigência Inicial', 'vigencia_inicial'),
        ('Vigência Final', 'vigencia_final'),
        ('Itens', 'total_itens'),
        ('Saldos', 'total_saldos'),
        ('Objeto', 'objeto'),
    ]
    _write_header(ws, cols)
    for r, row in enumerate(arps, 2):
        for c, (_, key) in enumerate(cols, 1):
            ws.cell(row=r, column=c, value=row.get(key, ''))
    _auto_width(ws)


def _sheet_itens_xlsx(wb, itens):
    ws = wb.create_sheet('Itens')
    cols = [
        ('ATA', 'numero_ata'),
        ('Item', 'numero_item'),
        ('Descrição', 'descricao'),
        ('Fornecedor NI', 'fornecedor_ni'),
        ('Fornecedor Nome', 'fornecedor_nome'),
        ('Valor Unitário', 'valor_unitario'),
        ('Valor Total', 'valor_total'),
        ('Qtd. Registrada', 'qtd_registrada'),
        ('Qtd. Homologada', 'qtd_homologada'),
        ('Qtd. Empenhada', 'qtd_empenhada_total'),
        ('% Empenhado', 'pct_empenhado'),
        ('Saldos', 'total_saldos'),
    ]
    _write_header(ws, cols)
    num_cols = {'valor_unitario', 'valor_total', 'qtd_registrada',
                'qtd_homologada', 'qtd_empenhada_total', 'pct_empenhado', 'total_saldos'}
    for r, row in enumerate(itens, 2):
        for c, (_, key) in enumerate(cols, 1):
            v = row.get(key)
            cell = ws.cell(row=r, column=c)
            if key in num_cols and v is not None:
                cell.value = float(v)
                if key in ('valor_unitario', 'valor_total'):
                    cell.number_format = 'R$ #,##0.00'
                elif key == 'pct_empenhado':
                    cell.number_format = '0.0"%"'
            else:
                cell.value = _v(v)
    _auto_width(ws)


def _sheet_saldos_xlsx(wb, saldos):
    ws = wb.create_sheet('Saldos por Unidade')
    cols = [
        ('ATA', 'numero_ata'),
        ('Item', 'numero_item'),
        ('Descrição Item', 'descricao_item'),
        ('Unidade', 'unidade'),
        ('Tipo', 'tipo'),
        ('Qtd. Registrada', 'qtd_registrada'),
        ('Qtd. Empenhada', 'qtd_empenhada'),
        ('Saldo Empenho', 'saldo_empenho'),
        ('Data Inclusão', 'data_inclusao'),
        ('Data Atualização', 'data_atualizacao'),
    ]
    _write_header(ws, cols)
    num_cols = {'qtd_registrada', 'qtd_empenhada', 'saldo_empenho'}
    for r, row in enumerate(saldos, 2):
        for c, (_, key) in enumerate(cols, 1):
            v = row.get(key)
            cell = ws.cell(row=r, column=c)
            if key in num_cols and v is not None:
                try:
                    cell.value = float(v)
                except Exception:
                    cell.value = _v(v)
            else:
                cell.value = _v(v)
    _auto_width(ws)


def render_arp_xlsx(bundle):
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_resumo_xlsx(wb, bundle['meta'])
    _sheet_arps_xlsx(wb, bundle['arps'])
    _sheet_itens_xlsx(wb, bundle['itens'])
    _sheet_saldos_xlsx(wb, bundle['saldos'])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CSV ────────────────────────────────────────────────────────────────────────

def render_arp_csv(bundle):
    """Flat CSV: one row per saldo, with ATA + item info repeated."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator='\n')
    writer.writerow([
        'ATA', 'Item', 'Descrição Item',
        'Fornecedor NI', 'Fornecedor Nome',
        'Valor Unitário', 'Valor Total',
        'Qtd. Registrada (Item)', 'Qtd. Homologada',
        'Qtd. Empenhada Total', '% Empenhado',
        'Unidade', 'Tipo',
        'Qtd. Registrada (Saldo)', 'Qtd. Empenhada', 'Saldo Empenho',
        'Data Inclusão', 'Data Atualização',
    ])
    # Build item lookup for denorm
    item_map = {(it['numero_ata'], it['numero_item']): it for it in bundle['itens']}
    for s in bundle['saldos']:
        it = item_map.get((s['numero_ata'], s['numero_item']), {})
        writer.writerow([
            s['numero_ata'], s['numero_item'], s['descricao_item'],
            it.get('fornecedor_ni', ''), it.get('fornecedor_nome', ''),
            _brl(it.get('valor_unitario')), _brl(it.get('valor_total')),
            _v(it.get('qtd_registrada')), _v(it.get('qtd_homologada')),
            _v(it.get('qtd_empenhada_total')), _pct(it.get('pct_empenhado')),
            s['unidade'], s['tipo'],
            _v(s['qtd_registrada']), _v(s['qtd_empenhada']), _v(s['saldo_empenho']),
            s['data_inclusao'], s['data_atualizacao'],
        ])
    return buf.getvalue().encode('utf-8')


# ── ODS ────────────────────────────────────────────────────────────────────────

def _ods_cell(value, bold=False):
    tc = TableCell()
    if value is None or value == '':
        tc.setAttribute('office:value-type', 'string')
    elif isinstance(value, (int, float)):
        tc.setAttribute('office:value-type', 'float')
        tc.setAttribute('office:value', str(value))
    else:
        tc.setAttribute('office:value-type', 'string')
    p = P(text=str(value) if value is not None else '')
    if bold:
        style = Style(name='bold', family='text')
        style.addElement(TextProperties(fontweight='bold'))
        p.setAttribute('text:style-name', 'bold')
    tc.addElement(p)
    return tc


def _ods_sheet(doc, name, headers, rows_data, key_order):
    table = Table(name=name)
    doc.spreadsheet.addElement(table)
    # Header row
    hr = TableRow()
    table.addElement(hr)
    for h in headers:
        hr.addElement(_ods_cell(h, bold=True))
    # Data rows
    for row in rows_data:
        tr = TableRow()
        table.addElement(tr)
        for key in key_order:
            v = row.get(key)
            tr.addElement(_ods_cell(v))


def render_arp_ods(bundle):
    doc = OpenDocumentSpreadsheet()

    # Resumo
    t = Table(name='Resumo')
    doc.spreadsheet.addElement(t)
    meta = bundle['meta']
    for k, v in [
        ('Código UASG', meta['codigo_uasg']),
        ('Nome', meta['nome_uasg']),
        ('UF', meta['sigla_uf']),
        ('Município', meta['municipio']),
        ('CNPJ', meta['cnpj']),
        ('Gerado em', meta['generated_at']),
        ('Total ARPs', meta['total_arps']),
        ('Total Itens', meta['total_itens']),
        ('Total Saldos', meta['total_saldos']),
    ]:
        tr = TableRow()
        t.addElement(tr)
        tr.addElement(_ods_cell(k, bold=True))
        tr.addElement(_ods_cell(v))

    # ARPs
    _ods_sheet(doc, 'ARPs',
               ['Nº ATA', 'Nº Controle PNCP', 'Vigência Inicial', 'Vigência Final',
                'Itens', 'Saldos', 'Objeto'],
               bundle['arps'],
               ['numero_ata', 'numero_controle', 'vigencia_inicial', 'vigencia_final',
                'total_itens', 'total_saldos', 'objeto'])

    # Itens
    _ods_sheet(doc, 'Itens',
               ['ATA', 'Item', 'Descrição', 'Fornecedor NI', 'Fornecedor Nome',
                'Valor Unitário', 'Valor Total', 'Qtd. Registrada', 'Qtd. Homologada',
                'Qtd. Empenhada', '% Empenhado', 'Saldos'],
               bundle['itens'],
               ['numero_ata', 'numero_item', 'descricao', 'fornecedor_ni', 'fornecedor_nome',
                'valor_unitario', 'valor_total', 'qtd_registrada', 'qtd_homologada',
                'qtd_empenhada_total', 'pct_empenhado', 'total_saldos'])

    # Saldos
    _ods_sheet(doc, 'Saldos por Unidade',
               ['ATA', 'Item', 'Descrição Item', 'Unidade', 'Tipo',
                'Qtd. Registrada', 'Qtd. Empenhada', 'Saldo Empenho',
                'Data Inclusão', 'Data Atualização'],
               bundle['saldos'],
               ['numero_ata', 'numero_item', 'descricao_item', 'unidade', 'tipo',
                'qtd_registrada', 'qtd_empenhada', 'saldo_empenho',
                'data_inclusao', 'data_atualizacao'])

    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()
