"""Render export bundle to .xlsx with auto-sized columns."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.exports.parse import fmt_br_money


# ─── Style helpers ───────────────────────────────────────────────────────────

def _header_font():
    return Font(bold=True, color='ffffff', size=11)


def _header_fill():
    return PatternFill('solid', fgColor='2c3e50')


def _header_align():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def _header_border():
    thin = Side(style='thin', color='ffffff')
    return Border(bottom=thin)


def _fmt_val(v):
    if v is None:
        return ''
    return str(v)


def _fmt_num(v):
    if v is None or v == '':
        return ''
    try:
        return fmt_br_money(float(v))
    except Exception:
        return _fmt_val(v)


def _fmt_bool(v):
    if v is None or v == '':
        return ''
    return 'Sim' if v else 'Não'


def _auto_columns(ws, min_width=10, max_width=60):
    """Calculate column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    clean = str(cell.value).replace('<', '').replace('>', '')
                    length = len(clean)
                    if length > max_len:
                        max_len = length
            except Exception:
                pass
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ─── Resumo sheet ────────────────────────────────────────────────────────────

def _sheet_resumo(wb, meta, empresa):
    ws = wb.create_sheet("Resumo")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    # Title row
    ws['A1'] = "Relatório de Empenhos"
    ws['A1'].font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 22

    data = [
        ("Razão/Nome", meta.get('search_name', '')),
        ("CNPJ", meta.get('cnpj', '')),
        ("Filtro Pregão", meta.get('pregao_filter', '') or '(todos)'),
        ("Status da busca", meta.get('status', '')),
        ("Gerado em", meta.get('generated_at', '')),
        ("Empenhos", meta.get('count', 0)),
        ("Enriquecidos", meta.get('enriched_count', 0)),
        ("Valor total (R$)", _fmt_num(meta.get('total_valor', 0))),
    ]
    for i, (k, v) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    if empresa:
        row = len(data) + 3
        ws.cell(row=row, column=1, value="Sanções / Cadastros").font = Font(bold=True, size=14)
        row += 1
        sanctions = [
            ('sancionadoCEIS', 'Inidôneos (CEIS)'),
            ('sancionadoCNEP', 'Penalidades (CNEP)'),
            ('sancionadoCEPIM', 'Entidades Privadas (CEPIM)'),
            ('sancionadoCEAF', 'CEAF'),
            ('favorecidoDespesas', 'Favorecido Despesas'),
            ('possuiContratacao', 'Possui Contratação'),
            ('convenios', 'Convênios'),
            ('favorecidoTransferencias', 'Favorecido Transferências'),
            ('participanteLicitacao', 'Participante Licitação'),
            ('emitiuNFe', 'Emitiu NFe'),
            ('beneficiadoRenunciaFiscal', 'Beneficiado Renúncia Fiscal'),
            ('isentoImuneRenunciaFiscal', 'Isento/Imune Renúncia Fiscal'),
            ('habilitadoRenunciaFiscal', 'Habilitado Renúncia Fiscal'),
        ]
        for k, label in sanctions:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=_fmt_bool(empresa.get(k)))
            row += 1

    _auto_columns(ws)


# ─── Empenhos sheet ─────────────────────────────────────────────────────────

def _sheet_empenhos(wb, empenhos):
    ws = wb.create_sheet("Empenhos")

    cols = [
        ('Data', 'data', 'str'),
        ('Documento', 'documento', 'str'),
        ('Resumido', 'documentoResumido', 'str'),
        ('Valor (R$)', 'valor', 'num'),
        ('Fase', 'fase', 'str'),
        ('Espécie', 'especie', 'str'),
        ('Favorecido', 'favorecido', 'str'),
        ('Cód. Favorecido', 'codigoFavorecido', 'str'),
        ('Nome Favorecido', 'nomeFavorecido', 'str'),
        ('UF Favorecido', 'ufFavorecido', 'str'),
        ('Cód. UG', 'codigoUg', 'str'),
        ('UG', 'ug', 'str'),
        ('Cód. UO', 'codigoUo', 'str'),
        ('UO', 'uo', 'str'),
        ('Cód. Órgão', 'codigoOrgao', 'str'),
        ('Órgão', 'orgao', 'str'),
        ('Cód. Órgão Superior', 'codigoOrgaoSuperior', 'str'),
        ('Órgão Superior', 'orgaoSuperior', 'str'),
        ('Função', 'funcao', 'str'),
        ('Subfunção', 'subfuncao', 'str'),
        ('Programa', 'programa', 'str'),
        ('Ação', 'acao', 'str'),
        ('Subtítulo', 'subTitulo', 'str'),
        ('Localizador Gasto', 'localizadorGasto', 'str'),
        ('Categoria', 'categoria', 'str'),
        ('Grupo', 'grupo', 'str'),
        ('Elemento', 'elemento', 'str'),
        ('Modalidade', 'modalidade', 'str'),
        ('Nº Processo', 'numeroProcesso', 'str'),
        ('Plano Orçamentário', 'planoOrcamentario', 'str'),
        ('Observação', 'observacao', 'str'),
        ('Autor', 'autor', 'str'),
        ('Favorecido Intermediário', 'favorecidoIntermediario', 'str'),
        ('Enriquecido', 'enriched', 'bool'),
    ]

    # Write header row with styling
    for i, (label, _, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _header_align()
        cell.border = _header_border()

    for e in empenhos:
        row_idx = empenhos.index(e) + 2
        for col_idx, (_, key, kind) in enumerate(cols, 1):
            v = e.get(key, '')
            if kind == 'num':
                if v is not None and v != '':
                    cell = ws.cell(row=row_idx, column=col_idx, value=float(v))
                    cell.number_format = '#,##0.00'
                else:
                    ws.cell(row=row_idx, column=col_idx, value='')
            elif kind == 'bool':
                ws.cell(row=row_idx, column=col_idx, value=_fmt_bool(v))
            else:
                ws.cell(row=row_idx, column=col_idx, value=_fmt_val(v))

    _auto_columns(ws, min_width=12, max_width=50)


# ─── Itens sheet ─────────────────────────────────────────────────────────────

def _sheet_itens(wb, itens):
    ws = wb.create_sheet("Itens")

    cols = [
        ('Documento', 'documento', 'str'),
        ('Sequencial', 'sequencial', 'str'),
        ('Cód. Item', 'codigoItemEmpenho', 'str'),
        ('Descrição', 'descricao', 'str'),
        ('Cód. Subelemento', 'codigoSubelemento', 'str'),
        ('Subelemento', 'descricaoSubelemento', 'str'),
        ('Valor Atual (R$)', 'valorAtual_num', 'num'),
    ]

    for i, (label, _, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _header_align()
        cell.border = _header_border()

    for it in itens:
        row_idx = itens.index(it) + 2
        for col_idx, (_, key, kind) in enumerate(cols, 1):
            v = it.get(key, '')
            if kind == 'num':
                if v is not None and v != '':
                    cell = ws.cell(row=row_idx, column=col_idx, value=float(v))
                    cell.number_format = '#,##0.00'
                else:
                    ws.cell(row=row_idx, column=col_idx, value='')
            else:
                ws.cell(row=row_idx, column=col_idx, value=_fmt_val(v))

    _auto_columns(ws, min_width=12, max_width=50)


# ─── Histórico sheet ─────────────────────────────────────────────────────────

def _sheet_historico(wb, historico):
    ws = wb.create_sheet("Histórico")

    cols = [
        ('Documento', 'documento', 'str'),
        ('Sequencial', 'sequencial', 'str'),
        ('Data', 'data', 'str'),
        ('Operação', 'operacao', 'str'),
        ('Quantidade', 'quantidade_num', 'num'),
        ('Valor Unitário (R$)', 'valorUnitario_num', 'num'),
        ('Valor Total (R$)', 'valorTotal_num', 'num'),
    ]

    for i, (label, _, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _header_align()
        cell.border = _header_border()

    for h in historico:
        row_idx = historico.index(h) + 2
        for col_idx, (_, key, kind) in enumerate(cols, 1):
            v = h.get(key, '')
            if kind == 'num':
                if v is not None and v != '':
                    cell = ws.cell(row=row_idx, column=col_idx, value=float(v))
                    cell.number_format = '#,##0.00'
                else:
                    ws.cell(row=row_idx, column=col_idx, value='')
            else:
                ws.cell(row=row_idx, column=col_idx, value=_fmt_val(v))

    _auto_columns(ws, min_width=12, max_width=50)


# ─── Documentos Relacionados sheet ─────────────────────────────────────────

def _sheet_relacionados(wb, relacionados):
    ws = wb.create_sheet("Documentos Relacionados")

    cols = [
        ('Empenho Origem', 'documento_empenho', 'str'),
        ('Data', 'data', 'str'),
        ('Fase', 'fase', 'str'),
        ('Documento', 'documento', 'str'),
        ('Resumido', 'documentoResumido', 'str'),
        ('Espécie', 'especie', 'str'),
        ('Órgão Superior', 'orgaoSuperior', 'str'),
        ('Órgão Vinculado', 'orgaoVinculado', 'str'),
        ('Unidade Gestora', 'unidadeGestora', 'str'),
        ('Elemento Despesa', 'elementoDespesa', 'str'),
        ('Favorecido', 'favorecido', 'str'),
        ('Valor (R$)', 'valor_num', 'num'),
        ('Link Portal', 'url_portal', 'link'),
    ]

    for i, (label, _, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _header_align()
        cell.border = _header_border()

    for rel in relacionados:
        row_idx = relacionados.index(rel) + 2
        for col_idx, (_, key, kind) in enumerate(cols, 1):
            v = rel.get(key, '')
            if kind == 'num':
                if v is not None and v != '':
                    cell = ws.cell(row=row_idx, column=col_idx, value=float(v))
                    cell.number_format = '#,##0.00'
                else:
                    ws.cell(row=row_idx, column=col_idx, value='')
            elif kind == 'link':
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                if v:
                    cell.hyperlink = v
                    cell.font = Font(color='0563C1', underline='single')
            else:
                ws.cell(row=row_idx, column=col_idx, value=_fmt_val(v))

    _auto_columns(ws, min_width=12, max_width=50)


# ─── Empresa sheet ────────────────────────────────────────────────────────────

def _sheet_empresa(wb, empresa):
    if not empresa:
        return
    ws = wb.create_sheet("Empresa")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50

    fields = [
        ('cnpj', 'CNPJ'),
        ('razaoSocial', 'Razão Social'),
        ('nomeFantasia', 'Nome Fantasia'),
        ('favorecidoDespesas', 'Favorecido Despesas'),
        ('possuiContratacao', 'Possui Contratação'),
        ('convenios', 'Convênios'),
        ('favorecidoTransferencias', 'Favorecido Transferências'),
        ('sancionadoCEIS', 'Sancionado CEIS (Inidôneos)'),
        ('sancionadoCNEP', 'Sancionado CNEP (Penalidades)'),
        ('sancionadoCEPIM', 'Sancionado CEPIM (Entidades Privadas)'),
        ('sancionadoCEAF', 'Sancionado CEAF'),
        ('participanteLicitacao', 'Participante Licitação'),
        ('emitiuNFe', 'Emitiu NFe'),
        ('beneficiadoRenunciaFiscal', 'Beneficiado Renúncia Fiscal'),
        ('isentoImuneRenunciaFiscal', 'Isento/Imune Renúncia Fiscal'),
        ('habilitadoRenunciaFiscal', 'Habilitado Renúncia Fiscal'),
    ]

    for i, (key, label) in enumerate(fields, 1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        v = empresa.get(key)
        if isinstance(v, bool):
            ws.cell(row=i, column=2, value=_fmt_bool(v))
        else:
            ws.cell(row=i, column=2, value=_fmt_val(v))


# ─── Main ────────────────────────────────────────────────────────────────────

def render_xlsx(bundle):
    """Render ExportBundle dict to .xlsx bytes."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _sheet_resumo(wb, bundle['meta'], bundle.get('empresa') or {})
    _sheet_empenhos(wb, bundle['empenhos'])
    _sheet_itens(wb, bundle['itens'])
    _sheet_historico(wb, bundle['historico'])
    _sheet_relacionados(wb, bundle['relacionados'])
    _sheet_empresa(wb, bundle.get('empresa') or {})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()